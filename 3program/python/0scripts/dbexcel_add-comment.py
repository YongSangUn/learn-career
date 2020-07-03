# -*- coding:utf-8 -*-

'''
数据库巡检表格相关操作，确保安装 openpyxl 模块。
    $ pip install openpyxl
'''

import openpyxl
import time
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill


class DatabaseExcelOperate(object):
    '''数据库巡检表格相关操作

    可对类内进行功能扩展，目前只有添加批注和数据统计。

    Args:
        excelFile(str)      : 数据文件

    '''
    def __init__(self, excelFile):
        self.excelFile = excelFile
        self.dbExcel = openpyxl.load_workbook(filename=excelFile)
        self.ws = self.dbExcel.active

    def totalData(self):
        '''统计信息

        统计 Agent & Disk 异常信息，如果 [Disk 异常] 不为 0 ，就会提示需要处理。

        NOTE：
            FFFFFF00 为 黄色 RGB 值
            FFFF0000 为 红色 RGB 值
        '''
        agentErrorNum = 0
        diskErrorNum = 0
        for row in self.ws.iter_rows():
            for cell in row:
                cellColor = cell.fill.fgColor.rgb
                if cellColor == "FFFFFF00":
                    agentErrorNum += 1
                if cellColor == "FFFF0000":
                    diskErrorNum += 1
        print("\tAgent 异常数   ：" + str(agentErrorNum))
        print("\tDisk  异常数   : " + str(diskErrorNum))
        if diskErrorNum != 0:
            print("*** 有 " + str(diskErrorNum) + " 个 [Disk 异常] 未添加批注，请处理。 ***")
        elif diskErrorNum == 0:
            print("无异常。")

    def addComment(self, diskCell, diskCommandText):
        '''添加批注

        这一步只是添加一个单元格的批注，在下一个模块中可以批量添加。

        Args:
            diskCell(str)           : 表格的单元格坐标，对应 commentList 的 key
            diskCommandText(str)    : 批注信息，对应 commentList[key] 的 value

        '''
        diskComment = Comment(diskCommandText, 'NOC')
        diskComment.width = 200
        diskComment.height = 100
        self.ws[diskCell].comment = diskComment
        diskFill = PatternFill(patternType='solid', fgColor="FFC125")
        self.ws[diskCell].fill = diskFill

    def addCommentList(self, commentList):
        '''批量添加批注

        通过传入的字典，自动识别添加批注。

        Args:
            commentList(str)        : 批注列表

        '''
        for cell in commentList.keys():
            diskCell = cell
            diskCommandText = commentList[cell]
            self.addComment(diskCell, diskCommandText)
            if self.ws[diskCell].comment.text == diskCommandText:
                print("\t==> Success add " + diskCell + " : " +
                      diskCommandText)

    def saveFile(self):
        # 保存并覆盖原文件
        self.dbExcel.save(self.excelFile)
        print("文件已保存并覆盖源文件。")

    def saveAsFile(self, newExcelFilePath):
        # 另存为文件
        self.dbExcel.save(newExcelFilePath)
        print("\n文件已另存为 \"" + newExcelFilePath + "\".")


def startWork(commentList):
    # 只是流程做成一个函数，后面直接调用就可以了，
    # 如果需要修改流程，在此函数中修改即可。
    today = time.strftime("%Y%m%d", time.localtime())
    todayDatabaseExcel = r'\\192.168.9.30\\系统运维\\运维\\11巡检工作\\数据库巡检\\巡检报告\\数据库巡检-%s.xlsx' % today
    excel = DatabaseExcelOperate(todayDatabaseExcel)
    print("添加批注：")
    excel.addCommentList(commentList)
    print("检查异常：")
    excel.totalData()  # 在添加批注后在检查异常数。
    # excel.saveFile()                                 # 保存和覆盖源文件，建议另存为。
    excel.saveAsFile("数据库巡检-%s.xlsx" % today)  # 另存为文件，括号中为文件


if __name__ == "__main__":
    '''
    key 为 坐标，value 为批注。
    将 坐标 和 对应的批注 添加到字典 commentList 中即可，注意最后一个 value 的 逗号 无需添加。
    '''
    commentList = {
        "G6": "移动硬盘，非数据库文件导致。",
        "G26": "为16年的归档内容",
        "E36": "为D盘，最后更新时间为2019/12/17，待观察。",
        "F36": "为E盘，最后更新时间为2019/12/17，待观察。",
        "E37": "为D盘，几个大文件更新时间为2019-5-14，待观察。",
        "M37": "为L盘，该盘本身较小，只有300G，数据库文件疑似备份，待观察.",
        "L37": "这台机器的L,D,S,K盘分别存的是2015，2016,2018,2019年的EhaiGPS数据库归档数据，基本不会变化",
        "T36": "为S盘，最后修改时间为2019-12-17，待观察",
        "T37": "为S盘，最后修改时间为2019-5-14，疑似备份，待观察"
    }
    startWork(commentList)
